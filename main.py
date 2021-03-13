import requests
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup

try:
    book = load_workbook(filename='ligue1.xlsx')  # TODO excelin adi değiştir
except FileNotFoundError:
    book = Workbook()
sheet = book.active


def getsite(url):
    headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36'}
    result = requests.get(url, headers=headers)
    return result


def outermain():
    oldurl = f"https://uk.soccerway.com/a/block_competition_matches_summary?block_id=page_competition_1_block_competition_matches_summary_10&callback_params=%7B%22page%22%3A%2217%22%2C%22block_service_id%22%3A%22competition_summary_block_competitionmatchessummary%22%2C%22round_id%22%3A%2258178%22%2C%22outgroup%22%3A%22%22%2C%22view%22%3A%221%22%2C%22competition_id%22%3A%2216%22%7D&action=changePage&params=%7B%22page%22%3A0%7D"

    for yr in range(20, -1, -1):  # TODO normalde 20,-1,-1
        year = getYear(yr)
        code = getCode(year)

        url = urlChangerYearly(oldurl, code)
        sheet.append([(str(2000+yr) + "-" + str(2001+yr) + ":"), ""])
        try:
            mainloop(url)
        except:
            continue
        book.save('ligue1.xlsx')


def urlchanger_week(url, index):
    y = []
    for i in range(50):
        if url[-i] == '%':
            y.append(-i)
            if len(y)==2:
                return url[:y[1]] + (url[y[1]:y[1]+3]) + "-" + str(index) + url[y[0]:] #TODO eksiyi çıkar


def getall(url):

    result = getsite(url)
    soup = BeautifulSoup(result.text, 'lxml')
    team1, team2 = getTeams(soup)  # DONE
    halftime, fulltime = getScores(soup)  # DONE
    details = getDetails(soup)
    goalstats = getGoals(soup)
    coach1, coach2 = getCoaches(soup)
    referee = getRef(soup)
    print(team1, halftime, team2, fulltime, details, coach1, coach2, referee, str(goalstats).strip())
    sheet.append([str(team1), str(fulltime), str(team2), str(halftime), str(details), str(goalstats), str(coach1),
                  str(coach2), str(referee),  str(url)])
    book.save('ligue1.xlsx')



def prettylink(semiurl):
    urltolist = (semiurl.split('\\'))
    resultsemiurl = ''
    for item in urltolist:
        resultsemiurl += item
    return resultsemiurl


def getTeams(soup):
    teams = soup.find_all('a', {'class', 'team-title'})
    return teams[0].text, teams[1].text


def getScores(soup):
    scores = soup.find('h3', {'class', 'thick scoretime'})
    spans = scores.find_all('span')
    ht1 = spans[1].text

    ht = ""
    for ch in ht1:
        if ((48 <= ord(ch) <= 57) or ord(ch) == 45) and (ord(ch) != 38):  # sadece '-' ve sayı almak için
            ht += ch

    ft = str(spans[0].next_sibling).strip()
    return ht, ft


def getDetails(soup):
    details = soup.find('div', {'class', 'details'})
    details2 = ''
    detail = ''
    for item in details.findChildren(recursive=False):
        if item.text != '':
            details2 += (item.text + ' ')
    for item in details2.split('\n'):
        detail += item
    return detail


def getGoals(soup):
    z = soup.find("ul", {"class": "scorer-info"})
    goals = ""
    if z is not None:
        for li in z.find_all('li'):

            span = li.find_all('span', recursive=False)
            try:
                goals += ('Gol: ' + span[0].find('a').text.strip() + ' ' + str(
                    span[0].find('a').next_sibling).strip() + '\n')

            except AttributeError:
                goals += ('Gol: ' + span[2].find('a').text.strip() + ' ' + str(
                    span[2].find('a').next_sibling).strip() + '\n')

            try:
                goals += ('Dakika: ' + span[0].find("span", {"class": "minute"}).text.strip() + '\n')
            except AttributeError:
                goals += ('Dakika: ' + span[2].find("span", {"class": "minute"}).text.strip() + '\n')

            try:
                goals += ('(Asist: ' + span[0].find("span", {"class": "assist"}).find(
                    'a').text.strip() + ')' + '\n')
            except AttributeError:
                try:
                    goals += ('(Asist: ' + span[2].find("span", {"class": "assist"}).find(
                        'a').text.strip() + ')' + '\n')
                except AttributeError:
                    pass

            goals += ('Skor: ' + span[1].text + '\n\n')
    return goals[:-2]


def getCoaches(soup):
    coach1, coach2 = '', ''
    try:
        coach1 = soup.find('div', {'class', 'combined-lineups-container'}).find_all('tbody')[0].find_all('tr', recursive=False)[-1].find('a').text
    except:
        pass
    try:
        coach2 = soup.find('div', {'class', 'combined-lineups-container'}).find_all('tbody')[1].find_all('tr', recursive=False)[-1].find('a').text
    except:
        pass
    return coach1, coach2


def getRef(soup):
    try:
        ref = soup.find(text='Referee:').parent.parent.find_all('td')[-1].text
        return ref
    except AttributeError:
        return ""


def getYear(year):
    if year >= 12:
        return "20"+str(year)+"20"+str(year+1)
    else:
        if (len(str(year)) == 1) and (len(str(year+1)) == 1):
            return "200"+str(year)+"-200"+str(year+1)
        elif (len(str(year)) == 1) and (len(str(year+1)) == 2):
            return "200" + str(year) + "-20" + str(year + 1)
        else:
            return "20" + str(year) + "-20" + str(year + 1)


def getCode(year):  # TODO liglerin linkini değiştir
    reulst = getsite('https://uk.soccerway.com/national/france/ligue-1/'+str(year))
    return reulst.url.split('/')[-2][1:]


def urlChangerYearly(url, code):
    a = (url.split('round_id')[1].split('%')[3])
    b = a[:2]
    b += code
    newurl = url.replace(a, b)
    return newurl


def mainloop(url):
    for x in range(0, 42):  # TODO normalde 0,50
        urlnew = urlchanger_week(url, x)

        result = getsite(urlnew)
        if result.status_code == 404:
            return urlchanger_week(urlnew, 0)

        soup = BeautifulSoup(result.text, 'lxml')
        tbody = soup.find('tbody')
        divs = tbody.find_all("div")

        print("week " + str(x + 1))

        for div in divs:
            links = div.find_all('a')[1]['href']
            link = ("https://uk.soccerway.com" + prettylink(links[2:-1]))
            print(link)
            getall(link)


if __name__ == "__main__":
    outermain()
