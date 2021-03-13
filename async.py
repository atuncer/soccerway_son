import asyncio
import aiohttp
import requests
from bs4 import BeautifulSoup
from collections import OrderedDict
from openpyxl import Workbook, load_workbook
import time

try:
    book = load_workbook(filename='prem.xlsx')  # TODO excelin adi değiştir
except FileNotFoundError:
    book = Workbook()
sheet = book.active

a = {}

URLS = []
# ['59136', '53145', '48730', '41547', '35992', '31553', '25191', '21322', '17950', '14829', '11930', '8926', '6786', '5054', '3570', '2150', '896', '29', '27', '26', '31']
codes = ['6786', '5054', '3570', '2150']


def getlinks():
    for j in codes:
        print("j is: ", codes.index(j))
        for i in range(45):
            print(i, end='')
            url1 = f"https://uk.soccerway.com/a/block_competition_matches_summary?block_id=page_competition_1_block_competition_matches_summary_10&callback_params=%7B%22page%22%3A%221%22%2C%22block_service_id%22%3A%22competition_summary_block_competitionmatchessummary%22%2C%22round_id%22%3A%22{j}%22%2C%22outgroup%22%3A%22%22%2C%22view%22%3A%221%22%2C%22competition_id%22%3A%227%22%7D&action=changePage&params=%7B%22page%22%3A{i}%7D"
            headers = {
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36'}
            urlx = requests.get(url1, headers=headers)

            url = urlx.text

            soup = BeautifulSoup(url, 'lxml')
            tbody = soup.find('tbody')
            # FT = soup.find('span', {'class', 'extra_time_score'})
            if tbody is None:
                break
            divs = tbody.find_all("div")
            b = []
            for div in divs:
                links = div.find_all('a')[1]['href']
                link = ("https://uk.soccerway.com" + prettylink(links[2:-1]))
                URLS.append(link)


async def main():
    getlinks()
    async with aiohttp.ClientSession() as session:
        await asyncio.wait([fetch(session, url, URLS.index(url)) for url in URLS])


async def fetch(session: aiohttp.ClientSession, url: str, n: int):
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36'}
    async with session.get(url, headers=headers) as result:
        print(f"Requested: {url}")
        txt = await result.text()
        soup = BeautifulSoup(txt, 'lxml')
        await asyncio.sleep(1)
        teams = soup.find_all('a', {'class', 'team-title'})

        team1, team2 = getTeams(soup)  # DONE
        halftime, fulltime = getScores(soup)  # DONE
        details = getDetails(soup)
        goalstats = getGoals(soup)
        coach1, coach2 = getCoaches(soup)
        referee = getRef(soup)

        a[n] = [str(team1), str(fulltime), str(team2), str(halftime), str(details), str(goalstats), str(coach1),
                  str(coach2), str(referee),  str(url)]

        # book.save('prem00-12.xlsx')  # TODO excelin adi değiştir


def prettylink(semiurl):
    urltolist = (semiurl.split('\\'))
    resultsemiurl = ''
    for item in urltolist:
        resultsemiurl += item
    return resultsemiurl


def getTeams(soup):
    try:
        teams = soup.find_all('a', {'class', 'team-title'})
        return teams[0].text, teams[1].text
    except:
        return 0, 0


def getScores(soup):
    scores = soup.find('h3', {'class', 'thick scoretime'})
    try:
        spans = scores.find_all('span')
        ht1 = spans[1].text

        ht = ""
        for ch in ht1:
            if ((48 <= ord(ch) <= 57) or ord(ch) == 45) and (ord(ch) != 38):  # sadece '-' ve sayı almak için
                ht += ch

        ft = str(spans[0].next_sibling).strip()
        return ht, ft
    except:
        return 0, 0


def getDetails(soup):
    try:
        details = soup.find('div', {'class', 'details'})
        details2 = ''
        detail = ''
        for item in details.findChildren(recursive=False):
            if item.text != '':
                details2 += (item.text + ' ')
        for item in details2.split('\n'):
            detail += item
        return detail
    except:
        return 0


def getGoals(soup):
    try:
        z = soup.find("ul", {"class": "scorer-info"})
        goals = ""
        if z is not None:
            for li in z.find_all('li'):

                span = li.find_all('span', recursive=False)
                try:
                    goals += ('Gol: ' + span[0].find('a').text.strip() + ' ' + str(
                        span[0].find('a').next_sibling).strip() + '\n')

                except AttributeError:
                    goals += ('Gol: ' + span[2].find('a').text.strip() + ' ' + str(
                        span[2].find('a').next_sibling).strip() + '\n')

                try:
                    goals += ('Dakika: ' + span[0].find("span", {"class": "minute"}).text.strip() + '\n')
                except AttributeError:
                    goals += ('Dakika: ' + span[2].find("span", {"class": "minute"}).text.strip() + '\n')

                try:
                    goals += ('(Asist: ' + span[0].find("span", {"class": "assist"}).find(
                        'a').text.strip() + ')' + '\n')
                except AttributeError:
                    try:
                        goals += ('(Asist: ' + span[2].find("span", {"class": "assist"}).find(
                            'a').text.strip() + ')' + '\n')
                    except AttributeError:
                        pass

                goals += ('Skor: ' + span[1].text + '\n\n')
        return goals[:-2]
    except:
        return 0


def getCoaches(soup):
    coach1, coach2 = '', ''
    coach1done = False
    coach2done = False
    try:
        if str(soup.find('div', {'class', 'combined-lineups-container'}).find_all('tbody')[0].find_all('tr',recursive=False)[-1].find('a')['href']).split('/')[2] == 'eric-hely':
            coach1 = 'E. Hely'
            coach1done = True
    except:
        pass

    try:
        if str(soup.find('div', {'class', 'combined-lineups-container'}).find_all('tbody')[1].find_all('tr',recursive=False)[-1].find('a')['href']).split('/')[2] == 'eric-hely':
            coach1 = 'E. Hely'
            coach2done = True
    except:
        pass

    try:
        if not coach1done:
            coach1 = soup.find('div', {'class', 'combined-lineups-container'}).find_all('tbody')[0].find_all('tr',recursive=False)[
                    -1].find('a').text
    except:
        pass
    try:
        if not coach2done:
            coach2 =  soup.find('div', {'class', 'combined-lineups-container'}).find_all('tbody')[1].find_all('tr', recursive=False)[
                    -1].find('a').text
    except:
        pass
    return coach1, coach2


def getRef(soup):
    try:
        ref = soup.find(text='Referee:').parent.parent.find_all('td')[-1].text
        return ref
    except:
        return 0


def saver():
    try:
        book.save('prem.xlsx')
    except PermissionError:
        time.sleep(20)
        saver()


if __name__ == '__main__':

    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    loop = asyncio.get_event_loop()
    loop.run_until_complete(main())

    a = OrderedDict(sorted(a.items()))
    for i in range(len(a)):
        try:
            sheet.append(a[i])
        except KeyError:
            sheet.append(['error', ''])
            continue
    saver()
